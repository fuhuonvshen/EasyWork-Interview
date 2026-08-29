"""File preview utilities — read Excel/CSV files and return text previews."""

import os

_MAX_PREVIEW_CHARS = 30_000


def _preview_dataframe(df, label: str, total_rows: int | None = None) -> str:
    """Format dataframe preview, truncating columns if too wide or long.

    Shows first 15 + last 5 columns when > 20 cols, and caps total output.
    `total_rows` is the real row count of the source file (may differ from
    `len(df)` when the preview only read the first N rows).
    """
    _MAX_COL_SHOW = 20
    _COL_SPLIT = 15  # show first 15, then last (20-15)=5

    ncols = len(df.columns)
    if total_rows is not None and total_rows > len(df):
        info = f"{label}: 共 {total_rows} 行, {ncols} 列（预览仅显示前 {len(df)} 行）\n"
    else:
        info = f"{label}: 预览 {len(df)} 行, {ncols} 列\n"

    if ncols > _MAX_COL_SHOW:
        front = list(df.columns[:_COL_SPLIT])
        back = list(df.columns[-(_MAX_COL_SHOW - _COL_SPLIT):])
        info += f"列名 (前{len(front)} + 后{len(back)}): {', '.join(str(c) for c in front)} ... {', '.join(str(c) for c in back)}\n\n"
        preview = df[front + back].head(10).to_string()
    else:
        info += f"列名: {', '.join(str(c) for c in df.columns)}\n\n"
        preview = df.head(10).to_string()

    if len(info) + len(preview) > _MAX_PREVIEW_CHARS:
        allowed = max(0, _MAX_PREVIEW_CHARS - len(info) - 100)
        preview = preview[:allowed] + f"\n\n... (预览已截断，共 {ncols} 列)"

    return info + preview


def read_excel_preview(file_path: str) -> str:
    """Return a text preview of Excel (.xlsx/.xls/.xlsm) with column names and first rows."""
    import pandas as pd
    try:
        if os.path.getsize(file_path) == 0:
            return "[空文件]"
        df = pd.read_excel(file_path, nrows=20)
        if df.empty:
            return "[空文件]"
        return _preview_dataframe(df, file_path, _excel_total_rows(file_path))
    except Exception as e:
        return f"读取 Excel 文件失败: {e}"


def _excel_total_rows(file_path: str) -> int | None:
    """Fast real row count of the first sheet (openpyxl max_row, no data read).
    Returns None for legacy .xls or on failure — preview falls back gracefully.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            n = ws.max_row
        finally:
            wb.close()
        return n
    except Exception:
        return None


def read_csv_preview(file_path: str, filename: str = "") -> str:
    """Return a text preview of a CSV file with inferred structure.
    Tries UTF-8 first, falls back to GBK for Chinese legacy systems.
    """
    import pandas as pd

    def _try_read(encoding: str):
        return pd.read_csv(file_path, nrows=20, encoding=encoding)

    encodings = ["utf-8", "gbk", "gb2312", "utf-16"]
    df = None
    for enc in encodings:
        try:
            df = _try_read(enc)
            break
        except UnicodeError:
            continue
    if not os.path.getsize(file_path):
        return "[空文件]"

    if df is None:
        return f"读取 CSV 文件失败: 无法以 utf-8/gbk/gb2312 编码解码，请确认文件格式"

    if df.empty:
        return "[空文件]"

    return _preview_dataframe(df, filename or file_path)
